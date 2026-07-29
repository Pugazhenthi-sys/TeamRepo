from openpyxl import Workbook
from openpyxl.styles import Font
from database import cursor
from datetime import datetime
import os


def generate_income_expense_report():

    os.makedirs("reports", exist_ok=True)

    wb = Workbook()

    ws = wb.active

    ws.title = "Income & Expenses"

    bold = Font(
        bold=True,
        size=12
    )

    row_num = 1

    # =====================================
    # TITLE
    # =====================================

    ws.cell(
        row=row_num,
        column=1,
        value="INCOME & EXPENSES REPORT"
    )

    ws.cell(
        row=row_num,
        column=1
    ).font = Font(
        bold=True,
        size=16
    )

    row_num += 3
    ws.cell(
        row=row_num,
        column=1,
        value=f"Generated On : {datetime.now().strftime('%d-%m-%Y %I:%M %p')}"
    )

    row_num += 3

    # =====================================
    # FETCH INCOME
    # =====================================

    cursor.execute("""
    SELECT
        bill_no,
        customer_name,
        bill_amount,
        paid_amount,
        profit,
        status
    FROM bill_status
    """)

    income_rows = cursor.fetchall()
    # =====================================
    # FETCH EXPENSES
    # =====================================

    cursor.execute("""
    SELECT
        date,
        description,
        amount
    FROM expenses
    """)

    expense_rows = cursor.fetchall()

    # =====================================
    # COMBINE DATA
    # =====================================

    records = []

    for (
        bill_no,
        customer_name,
        bill_amount,
        paid_amount,
        profit,
        status
    ) in income_rows:

        if status != "Completed":
            continue

        earned_profit = profit

        cursor.execute("""
        SELECT date
        FROM sales_register
        WHERE bill_no=?
        """, (bill_no,))

        row = cursor.fetchone()

        if not row:
            continue

        bill_date = row[0]

        records.append(
            (
                bill_date,
                "Income",
                f"Bill No {bill_no}",
                round(earned_profit, 2)
            )
        )

    for date, description, amount in expense_rows:

        records.append(
            (
                date,
                "Expense",
                description,
                amount
            )
        )

    records = sorted(
        records,
        key=lambda x: datetime.strptime(
            x[0],
            "%d-%m-%Y"
        )
    )

    current_month = None
    current_year = None

    month_income = 0
    month_expense = 0

    year_income = 0
    year_expense = 0

    total_income = 0
    total_expense = 0

    # =====================================
    # LOOP
    # =====================================

    for date, record_type, description, amount in records:

        dt = datetime.strptime(
            date,
            "%d-%m-%Y"
        )

        month_name = dt.strftime(
            "%B %Y"
        ).upper()

        year_name = dt.strftime(
            "%Y"
        )

        # =================================
        # YEAR TOTAL
        # =================================

        if current_year and year_name != current_year:

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value=f"YEAR TOTAL - {current_year}"
            ).font = bold

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value="Income"
            ).font = bold

            ws.cell(
                row=row_num,
                column=2,
                value=year_income
            )
            

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value="Expenses"
            ).font = bold

            ws.cell(
                row=row_num,
                column=2,
                value=year_expense
            )
            

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value="Net Income"
            ).font = bold

            ws.cell(
                row=row_num,
                column=2,
                value=year_income - year_expense
            )
            

            row_num += 3

            year_income = 0
            year_expense = 0

        # =================================
        # MONTH TOTAL
        # =================================

        if current_month and month_name != current_month:

            ws.cell(
                row=row_num,
                column=1,
                value=f"MONTH TOTAL - {current_month}"
            ).font = bold

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value="Income"
            ).font = bold

            ws.cell(
                row=row_num,
                column=2,
                value=month_income
            )
            

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value="Expenses"
            ).font = bold

            ws.cell(
                row=row_num,
                column=2,
                value=month_expense
            )
            

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value="Net Income"
            ).font = bold

            ws.cell(
                row=row_num,
                column=2,
                value=month_income - month_expense
            )
            

            row_num += 3

            month_income = 0
            month_expense = 0

        # =================================
        # MONTH HEADER
        # =================================

        if month_name != current_month:

            current_month = month_name

            if current_year != year_name:
                current_year = year_name

            ws.cell(
                row=row_num,
                column=1,
                value=month_name
            ).font = Font(
                bold=True,
                size=14
            )

            row_num += 1

            headers = [
                "Date",
                "Type",
                "Description",
                "Amount"
            ]

            for col, head in enumerate(
                headers,
                start=1
            ):

                ws.cell(
                    row=row_num,
                    column=col,
                    value=head
                ).font = bold

            row_num += 1

        # =================================
        # DATA ROW
        # =================================

        ws.cell(
            row=row_num,
            column=1,
            value=date
        )

        ws.cell(
            row=row_num,
            column=2,
            value=record_type
        )

        ws.cell(
            row=row_num,
            column=3,
            value=description
        )

        ws.cell(
            row=row_num,
            column=4,
            value=amount
        )
        

        row_num += 1

        # =================================
        # TOTALS
        # =================================

        if record_type == "Income":

            month_income += amount
            year_income += amount
            total_income += amount

        else:

            month_expense += amount
            year_expense += amount
            total_expense += amount

    # =====================================
    # LAST MONTH
    # =====================================

    if records:

        ws.cell(
            row=row_num,
            column=1,
            value=f"MONTH TOTAL - {current_month}"
        ).font = bold

        row_num += 1

        ws.cell(row=row_num, column=1, value="Income").font = bold
        ws.cell(row=row_num, column=2, value=month_income)

        row_num += 1

        ws.cell(row=row_num, column=1, value="Expenses").font = bold
        ws.cell(row=row_num, column=2, value=month_expense)

        row_num += 1

        ws.cell(row=row_num, column=1, value="Net Income").font = bold
        ws.cell(row=row_num, column=2, value=month_income - month_expense)

        row_num += 3

        # =================================
        # LAST YEAR
        # =================================

        ws.cell(
            row=row_num,
            column=1,
            value=f"YEAR TOTAL - {current_year}"
        ).font = bold

        row_num += 1

        ws.cell(row=row_num, column=1, value="Income").font = bold
        ws.cell(row=row_num, column=2, value=year_income)

        row_num += 1

        ws.cell(row=row_num, column=1, value="Expenses").font = bold
        ws.cell(row=row_num, column=2, value=year_expense)

        row_num += 1

        ws.cell(row=row_num, column=1, value="Net Income").font = bold
        ws.cell(row=row_num, column=2, value=year_income - year_expense)

        row_num += 4

    # =====================================
    # OVERALL SUMMARY
    # =====================================

    ws.cell(
        row=row_num,
        column=1,
        value="OVERALL INCOME & EXPENSES SUMMARY"
    ).font = Font(
        bold=True,
        size=14
    )

    row_num += 2

    summary = [
        ["Total Income", total_income],
        ["Total Expenses", total_expense],
        ["Net Income", total_income - total_expense]
    ]

    for item in summary:

        ws.cell(
            row=row_num,
            column=1,
            value=item[0]
        ).font = bold

        ws.cell(
            row=row_num,
            column=2,
            value=item[1]
        )
        

        row_num += 1


    # =====================================
    # PROFIT SUMMARY
    # =====================================

    row_num += 3

    ws.cell(
        row=row_num,
        column=1,
        value="PROFIT SUMMARY"
    ).font = Font(
        bold=True,
        size=14
    )

    row_num += 2

   
    # Overall Profit
    cursor.execute("""
    SELECT IFNULL(SUM(profit),0)
    FROM bill_status
    """)

    overall_profit = cursor.fetchone()[0]
    # Pending Profit
    cursor.execute("""
    SELECT IFNULL(SUM(profit),0)
    FROM bill_status
    WHERE status='Pending'
    """)

    pending_profit = cursor.fetchone()[0]
    realized_profit = overall_profit-pending_profit

    profit_summary = [
        ["Realized Profit", realized_profit],
        ["Pending Profit", pending_profit],
        ["Total Profit", realized_profit + pending_profit]
    ]

    for item in profit_summary:

        ws.cell(
            row=row_num,
            column=1,
            value=item[0]
        ).font = bold

        ws.cell(
            row=row_num,
            column=2,
            value=item[1]
        )
        

        row_num += 1

    # =====================================
    # AUTO WIDTH
    # =====================================

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_length + 5

    # =====================================
    # SAVE
    # =====================================

    file_path = os.path.abspath(
        os.path.join(
            "reports",
            "Income_Expenses_Report.xlsx"
        )
    )

    wb.save(file_path)

    return file_path
