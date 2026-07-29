# ==========================================
# billing.py
# COMPLETE FINAL VERSION
# ==========================================

import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image

from openpyxl.styles import (
    Alignment,
    Font,
    Border,
    Side
)

from database import conn, cursor

from ledger import (
    create_customer_if_not_exists,
    get_balance
)

from pdf_generator import (
    generate_pdf_bill
)
def get_next_global_bill_no():

    cursor.execute("""
    SELECT current_bill_no
    FROM bill_counter
    WHERE id=1
    """)

    current = cursor.fetchone()[0]

    next_no = current + 1

    cursor.execute("""
    UPDATE bill_counter
    SET current_bill_no=?
    WHERE id=1
    """, (next_no,))

    conn.commit()

    return next_no

# ==========================================
# CREATE CUSTOMER FOLDER
# ==========================================

def create_vendor_folder(vendor_name):

    base_dir = "data/vendors"

    vendor_path = os.path.join(
        base_dir,
        f"vendor_{vendor_name}"
    )

    invoices_path = os.path.join(
        vendor_path,
        "invoices",
        "generated"
    )

    os.makedirs(
        invoices_path,
        exist_ok=True
    )

    return invoices_path

# ==========================================
# NEXT BILL NUMBER
# ==========================================

def get_next_bill_number(
    path,
    vendor_name
):

    files = os.listdir(path)

    numbers = []

    for f in files:

        if (
            f.startswith(
                vendor_name.lower()
            )
            and
            f.endswith(".xlsx")
        ):

            try:

                num = int(
                    f.rsplit("_", 1)[1]
                    .split(".")[0]
                )

                numbers.append(num)

            except:
                pass

    return (
        max(numbers) + 1
        if numbers else 1
    )

# ==========================================
# GENERATE BILL GUI
# ==========================================

def generate_bill_gui(
    vendor_name,
    date,
    bags,
    bag_size,
    particular,
    rate,
    purchase_rate,
    duty_per_kg,
    vehicle_no
):

    # ======================================
    # CREATE CUSTOMER
    # ======================================

    customer_id = (
        create_customer_if_not_exists(
            vendor_name
        )
    )

    # ======================================
    # CREATE PATH
    # ======================================

    path = create_vendor_folder(
        vendor_name
    )

    # ======================================
    # BILL NUMBER
    # ======================================

    bill_no = get_next_global_bill_no()

    # ======================================
    # FILE NAME
    # ======================================

    file_name = (
        f"{vendor_name.lower()}_"
        f"{str(bill_no).zfill(3)}.xlsx"
    )

    file_path = os.path.join(
        path,
        file_name
    )

    # ======================================
    # CALCULATIONS
    # ======================================

    weight = bags * bag_size

    amount = weight * rate

    profit_per_kg = (
        rate - purchase_rate
    )

    profit = (
        weight * profit_per_kg
    )

    # ======================================
    # LORRY RENT
    # ======================================

    lorry_rent = (
        weight * duty_per_kg
    )

    # ======================================
    # FINAL BILL
    # ======================================

    final_bill = (
        amount + lorry_rent
    )

    # ======================================
    # PREVIOUS BALANCE
    # ======================================

    previous_balance = (
        get_balance(customer_id)
    )

    # ======================================
    # TOTAL BALANCE
    # ======================================

    total_balance = (
        previous_balance + final_bill
    )

    # ======================================
    # SAVE TRANSACTION
    # ======================================

    cursor.execute("""
    INSERT INTO transactions
    (
        customer_id,
        type,
        amount,
        description,
        date
    )
    VALUES (?, ?, ?, ?, ?)
    """, (
        customer_id,
        "credit",
        final_bill,
        f"Bill No {bill_no}",
        date
    ))

    conn.commit()

    # ======================================
    # SAVE PROFIT
    # ======================================

    cursor.execute("""
    INSERT INTO income
    (
        customer_name,
        bill_no,
        profit,
        date
    )
    VALUES (?, ?, ?, ?)
    """, (
        vendor_name,
        bill_no,
        profit,
        date
    ))

    conn.commit()
    # ======================================
    # SAVE SALES REGISTER
    # ======================================

    actual_cost = (
        weight * purchase_rate
    ) + lorry_rent

    cursor.execute("""
    INSERT INTO sales_register
    (
        customer_name,
        bill_no,
        particular,
        bags,
        bag_size,
        weight,
        purchase_rate,
        selling_rate,
        lorry_rent,
        actual_cost,
        selling_amount,
        profit,
        date
    )
    VALUES
    (
    ?,?,?,?,?,?,?,?,?,?,?,?,?
    )
    """,
    (
        vendor_name,
        bill_no,
        particular,
        bags,
        bag_size,
        weight,
        purchase_rate,
        rate,
        lorry_rent,
        actual_cost,
        final_bill,
        profit,
        date
    ))
    # ======================================
    # SAVE BILL STATUS
    # ======================================

    cursor.execute("""
    INSERT INTO bill_status
    (
        bill_no,
        customer_name,
        particular,
        bill_amount,
        paid_amount,
        profit,
        status
    )
    VALUES
    (
        ?, ?, ?, ?, ?, ?, ?
    )
    """,
    (
        bill_no,
        vendor_name,
        particular,
        final_bill,
        0,
        profit,
        "Pending"
    ))

    conn.commit()

    # ======================================
    # CREATE EXCEL
    # ======================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Bill"

    # ======================================
    # STYLES
    # ======================================

    bold = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ======================================
    # COLUMN WIDTH
    # ======================================

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    # ======================================
    # HEADER
    # ======================================

    ws.merge_cells("A1:E1")

    ws["A1"] = (
        "SHREE SAI SARAVANABHAVA TRADERS"
    )

    ws["A1"].font = Font(
        size=14,
        bold=True
    )

    ws["A1"].alignment = center

    ws.merge_cells("A2:E2")

    ws["A2"] = "Rice Merchants"

    ws["A2"].alignment = center

    ws.merge_cells("A3:E3")

    ws["A3"] = (
        "GSTIN: 33BFIPM8973G1Z1 | "
        "Phone: 7904434418"
    )

    ws["A3"].alignment = center

    # ======================================
    # BILL DETAILS
    # ======================================

    ws["A5"] = "Bill No:"
    ws["B5"] = bill_no

    ws["D5"] = "Date:"
    ws["E5"] = date

    ws["A6"] = "To:"
    ws["B6"] = vendor_name

    # ======================================
    # TABLE HEADER
    # ======================================

    headers = [
        "Bags",
        "Particular",
        "Qty (kg)",
        "Rate",
        "Amount"
    ]

    for col, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=8,
            column=col
        )

        cell.value = header

        cell.font = bold

        cell.alignment = center

        cell.border = border

    # ======================================
    # TABLE DATA
    # ======================================

    data = [
        f"{bags} ({bag_size}kg)",
        particular,
        weight,
        rate,
        amount
    ]

    for col, value in enumerate(
        data,
        start=1
    ):

        cell = ws.cell(
            row=9,
            column=col
        )

        cell.value = value

        cell.border = border

    # ======================================
    # BILL SUMMARY
    # ======================================

    ws["D11"] = "Previous Balance:"
    ws["E11"] = previous_balance

    ws["D12"] = "Lorry Rent:"
    ws["E12"] = lorry_rent

    ws["D13"] = "Final Bill:"
    ws["E13"] = final_bill

    ws["D14"] = "Total Balance:"
    ws["E14"] = total_balance

    # ======================================
    # BANK DETAILS
    # ======================================

    ws["A15"] = (
        "A/c No: 614305008545"
    )

    ws["A16"] = (
        "IFSC: ICICI0006143"
    )

    ws["A17"] = (
        "BRANCH: ICICI Bank, VELLORE"
    )

    # ======================================
    # VEHICLE NUMBER
    # ======================================

    ws["A19"] = "VEHICLE NO:"
    ws["B19"] = vehicle_no

    # ======================================
    # SIGNATURE
    # ======================================

    ws["C15"] = "Authorized Signature:"

    ws["D15"] = "Mathivanan"

    ws["C15"].font = bold

    ws["D15"].font = bold
    # Load signature image
    signature = Image("signature.png")

    # Resize if needed (optional)
    signature.width = 120
    signature.height = 50

    # Place signature near D15 (adjust anchor as needed)
    ws.add_image(signature, "D16")
    # ======================================
    # SAVE EXCEL
    # ======================================

    wb.save(file_path)

    # ======================================
    # GENERATE PDF
    # ======================================

    pdf_file = generate_pdf_bill(
        vendor_name,
        bill_no,
        date,
        bags,
        bag_size,
        particular,
        weight,
        rate,
        amount,
        lorry_rent,
        final_bill,
        total_balance,
        vehicle_no
    )

    # ======================================
    # RETURN
    # ======================================

    return (
        file_path,
        pdf_file,
        total_balance,
        profit
    )