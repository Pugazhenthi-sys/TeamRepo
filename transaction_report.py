from openpyxl import Workbook
from openpyxl.styles import Font
from database import cursor
from datetime import datetime
import os


def generate_transaction_report():

    os.makedirs("reports", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Business Report"

    bold = Font(bold=True, size=12)

    row_num = 1

    # =====================================
    # TITLE
    # =====================================

    ws.cell(row=row_num, column=1, value="BUSINESS TRANSACTION REPORT")
    ws.cell(row=row_num, column=1).font = Font(
        bold=True,
        size=16
    )

    row_num += 3

    # =====================================
    # FETCH DATA
    # =====================================

    cursor.execute("""
    SELECT
        bill_no,
        date,
        customer_name,
        particular,
        bags,
        weight,
        actual_cost,
        selling_amount,
        profit
    FROM sales_register
    """)

    rows = cursor.fetchall()

    # =====================================
    # SORT BY DATE
    # =====================================

    rows = sorted(
        rows,
        key=lambda x: datetime.strptime(
            x[1],
            "%d-%m-%Y"
        )
    )

    current_month = None
    current_year = None

    # =====================================
    # OVERALL TOTALS
    # =====================================

    overall_bills = 0
    overall_bags = 0
    overall_weight = 0
    overall_cost = 0
    overall_sales = 0
    overall_profit = 0

    # =====================================
    # MONTH TOTALS
    # =====================================

    month_bills = 0
    month_bags = 0
    month_weight = 0
    month_cost = 0
    month_sales = 0
    month_profit = 0

    # =====================================
    # YEAR TOTALS
    # =====================================

    year_bills = 0
    year_bags = 0
    year_weight = 0
    year_cost = 0
    year_sales = 0
    year_profit = 0

    # =====================================
    # LOOP DATA
    # =====================================

    for data in rows:

        (
            bill_no,
            bill_date,
            customer,
            particular,
            bags,
            weight,
            cost,
            sales,
            profit
        ) = data

        dt = datetime.strptime(
            bill_date,
            "%d-%m-%Y"
        )

        month_name = dt.strftime("%B %Y").upper()
        year_name = dt.strftime("%Y")

        # =================================
        # NEW YEAR
        # =================================

        if current_year and year_name != current_year:

            row_num += 1

            ws.cell(
                row=row_num,
                column=1,
                value=f"YEAR TOTAL - {current_year}"
            ).font = bold

            row_num += 1

            headers = [
                "Bills",
                "Bags",
                "Weight",
                "Cost",
                "Sales",
                "Profit"
            ]

            for col, head in enumerate(headers, start=1):
                ws.cell(row=row_num, column=col, value=head).font = bold

            row_num += 1

            values = [
                year_bills,
                year_bags,
                year_weight,
                year_cost,
                year_sales,
                year_profit
            ]

            for col, val in enumerate(values, start=1):
                ws.cell(row=row_num, column=col, value=val)

            row_num += 3

            year_bills = 0
            year_bags = 0
            year_weight = 0
            year_cost = 0
            year_sales = 0
            year_profit = 0

        # =================================
        # NEW MONTH
        # =================================

        if current_month and month_name != current_month:

            ws.cell(
                row=row_num,
                column=1,
                value=f"MONTH TOTAL - {current_month}"
            ).font = bold

            row_num += 1

            headers = [
                "Bills",
                "Bags",
                "Weight",
                "Cost",
                "Sales",
                "Profit"
            ]

            for col, head in enumerate(headers, start=1):
                ws.cell(row=row_num, column=col, value=head).font = bold

            row_num += 1

            values = [
                month_bills,
                month_bags,
                month_weight,
                month_cost,
                month_sales,
                month_profit
            ]

            for col, val in enumerate(values, start=1):
                ws.cell(row=row_num, column=col, value=val)

            row_num += 3

            month_bills = 0
            month_bags = 0
            month_weight = 0
            month_cost = 0
            month_sales = 0
            month_profit = 0

        # =================================
        # MONTH HEADER
        # =================================

        if month_name != current_month:

            current_month = month_name

            if current_year != year_name:
                current_year = year_name

            ws.cell(
                row=row_num,
                column=1,
                value=month_name
            ).font = Font(
                bold=True,
                size=14
            )

            row_num += 1

            headers = [
                "Bill No",
                "Date",
                "Customer",
                "Product",
                "Bags",
                "Weight",
                "Cost",
                "Sales",
                "Profit"
            ]

            for col, head in enumerate(headers, start=1):
                ws.cell(
                    row=row_num,
                    column=col,
                    value=head
                ).font = bold

            row_num += 1

        # =================================
        # TRANSACTION ROW
        # =================================

        values = [
            bill_no,
            bill_date,
            customer,
            particular,
            bags,
            weight,
            cost,
            sales,
            profit
        ]

        for col, val in enumerate(values, start=1):
            ws.cell(
                row=row_num,
                column=col,
                value=val
            )

        row_num += 1

        # =================================
        # MONTH TOTALS
        # =================================

        month_bills += 1
        month_bags += bags or 0
        month_weight += weight or 0
        month_cost += cost or 0
        month_sales += sales or 0
        month_profit += profit or 0

        # =================================
        # YEAR TOTALS
        # =================================

        year_bills += 1
        year_bags += bags or 0
        year_weight += weight or 0
        year_cost += cost or 0
        year_sales += sales or 0
        year_profit += profit or 0

        # =================================
        # OVERALL TOTALS
        # =================================

        overall_bills += 1
        overall_bags += bags or 0
        overall_weight += weight or 0
        overall_cost += cost or 0
        overall_sales += sales or 0
        overall_profit += profit or 0

    # =====================================
    # LAST MONTH TOTAL
    # =====================================

    if rows:

        ws.cell(
            row=row_num,
            column=1,
            value=f"MONTH TOTAL - {current_month}"
        ).font = bold

        row_num += 1

        headers = [
            "Bills",
            "Bags",
            "Weight",
            "Cost",
            "Sales",
            "Profit"
        ]

        for col, head in enumerate(headers, start=1):
            ws.cell(row=row_num, column=col, value=head).font = bold

        row_num += 1

        values = [
            month_bills,
            month_bags,
            month_weight,
            month_cost,
            month_sales,
            month_profit
        ]

        for col, val in enumerate(values, start=1):
            ws.cell(row=row_num, column=col, value=val)

        row_num += 3

        # =================================
        # LAST YEAR TOTAL
        # =================================

        ws.cell(
            row=row_num,
            column=1,
            value=f"YEAR TOTAL - {current_year}"
        ).font = bold

        row_num += 1

        for col, head in enumerate(headers, start=1):
            ws.cell(row=row_num, column=col, value=head).font = bold

        row_num += 1

        values = [
            year_bills,
            year_bags,
            year_weight,
            year_cost,
            year_sales,
            year_profit
        ]

        for col, val in enumerate(values, start=1):
            ws.cell(row=row_num, column=col, value=val)

        row_num += 4

    # =====================================
    # OVERALL SUMMARY
    # =====================================

    ws.cell(
        row=row_num,
        column=1,
        value="OVERALL BUSINESS SUMMARY"
    ).font = Font(
        bold=True,
        size=14
    )

    row_num += 2

    cursor.execute("""
    SELECT COUNT(DISTINCT customer_name)
    FROM sales_register
    """)

    total_customers = cursor.fetchone()[0] or 0

    # =====================================
    # TOTAL PENDING BALANCE
    # =====================================

    cursor.execute("""
    SELECT id
    FROM customers
    """)

    customers = cursor.fetchall()

    total_pending = 0

    for customer in customers:

        customer_id = customer[0]

        cursor.execute("""
        SELECT type, amount
        FROM transactions
        WHERE customer_id=?
        """, (customer_id,))

        transactions = cursor.fetchall()

        balance = 0

        for t in transactions:

            if t[0] == "credit":
                balance += t[1]
            else:
                balance -= t[1]

        if balance > 0:
            total_pending += balance

    # =====================================
    # PROFIT BREAKUP
    # =====================================

    

    cursor.execute("""
    SELECT IFNULL(SUM(profit),0)
    FROM bill_status
    WHERE status='Pending'
    """)
    pending_profit = cursor.fetchone()[0]
    realized_profit = overall_profit - pending_profit

    summary = [
        ["Total Bills", overall_bills],
        ["Total Customers", total_customers],
        ["Total Bags", overall_bags],
        ["Total Weight", overall_weight],
        ["Total Sales", overall_sales],

        ["Realized Profit", realized_profit],
        ["Pending Profit", pending_profit],

        ["Total Profit", overall_profit],

        ["Overall Outstanding", total_pending]
    ]

    for item in summary:
        ws.cell(row=row_num, column=1, value=item[0]).font = bold
        ws.cell(row=row_num, column=2, value=item[1])
        row_num += 1

    # =====================================
    # AUTO WIDTH
    # =====================================

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_length + 5

    # =====================================
    # SAVE
    # =====================================

    file_path = os.path.abspath(
        os.path.join(
            "reports",
            "Transaction_Report.xlsx"
        )
    )

    wb.save(file_path)

    return file_path